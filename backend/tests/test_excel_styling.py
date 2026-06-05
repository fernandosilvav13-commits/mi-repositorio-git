import pytest
from app.services.excel_service import ExcelService
from openpyxl import load_workbook
import os

@pytest.fixture
def excel_service():
    return ExcelService()

def test_excel_styling_unmatched_rows(excel_service):
    """
    Test that rows with unmatched cross-reference data are styled with YELLOW_FILL.
    Currently, the service only handles RED_FILL for critical failures.
    """
    columns = ["Nombre", "Email"]
    rows = [
        {"Nombre": "Juan", "Email": "juan@example.com"},
        {"Nombre": "Maria", "Email": "NO ENCONTRADO"} # Unmatched row
    ]
    
    file_path = excel_service.generate(columns, rows)
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Row 2 (Juan) - should have no fill
    assert ws.cell(row=2, column=2).fill.start_color.index == "00000000" or ws.cell(row=2, column=2).fill.fill_type is None
    
    # Row 3 (Maria) - should have YELLOW_FILL (FFFF44)
    # This test is expected to fail as the logic is not implemented
    assert ws.cell(row=3, column=2).fill.start_color.index == "00FFFF44"
    
    if os.path.exists(file_path):
        os.remove(file_path)

def test_excel_ref_prefixing(excel_service):
    """
    Test that cross-reference columns are prefixed with [REF].
    Currently, the service uses column names as-is.
    """
    columns = ["Nombre", "Email"]
    # We want to signify that "Email" comes from a crossref
    crossref_columns = ["Email"] 
    
    rows = [{"Nombre": "Juan", "Email": "juan@example.com"}]
    
    # Desired API change: pass crossref_columns or similar
    # This test is expected to fail or need adjustment
    file_path = excel_service.generate(columns, rows, crossref_columns=crossref_columns)
    wb = load_workbook(file_path)
    ws = wb.active
    
    assert ws.cell(row=1, column=1).value == "Nombre"
    assert ws.cell(row=1, column=2).value == "[REF] Email"
    
    if os.path.exists(file_path):
        os.remove(file_path)
