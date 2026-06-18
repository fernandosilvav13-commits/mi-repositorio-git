from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from app.core.config import settings

RED_FILL = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
GREEN_FILL = PatternFill(start_color="44FF44", end_color="44FF44", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF44", end_color="FFFF44", fill_type="solid")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NO_ENCONTRADO_TEXT = "NO ENCONTRADO"


class ExcelService:
    def __init__(self):
        self.output_dir = Path(settings.output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(
        self,
        columns: list[str],
        rows: list[dict],
        rules_triggered: list[list[dict]] | None = None,
        crossref_columns: list[str] | None = None,
    ) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"

        header_fill = PatternFill(
            start_color="2F5496", end_color="2F5496", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, col_name in enumerate(columns, 1):
            display_name = col_name
            if crossref_columns and col_name in crossref_columns:
                display_name = f"[REF] {col_name}"
            cell = ws.cell(row=1, column=col_idx, value=display_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row_data in enumerate(rows, 2):
            row_rules = rules_triggered[row_idx - 2] if rules_triggered else []

            has_no_encontrado = any(
                v == "NO ENCONTRADO" for v in row_data.values()
            )
            all_no_encontrado = has_no_encontrado and all(
                v == "NO ENCONTRADO" for v in row_data.values()
            )

            for col_idx, col_name in enumerate(columns, 1):
                value = row_data.get(col_name, NO_ENCONTRADO_TEXT)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                if all_no_encontrado:
                    cell.fill = RED_FILL
                elif has_no_encontrado:
                    cell.fill = YELLOW_FILL

            if all_no_encontrado:
                continue

            for rule in row_rules:
                action = rule.get("action", {})
                if action.get("type") == "fill_row":
                    color = action.get("params", {}).get("color", "")
                    fill = self._get_fill(color)
                    for col_idx in range(1, len(columns) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill

        for col_idx, col_name in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(
                12, len(col_name) * 2
            )

        output_path = self.output_dir / "resultado.xlsx"
        wb.save(str(output_path))
        return str(output_path)

    def _get_fill(self, color: str) -> PatternFill:
        match color.upper():
            case "GREEN":
                return GREEN_FILL
            case "YELLOW":
                return YELLOW_FILL
            case "RED":
                return RED_FILL
            case _:
                return PatternFill()
